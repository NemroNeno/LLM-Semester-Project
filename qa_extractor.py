# qa_extractor.py

import os
import re
from dotenv import load_dotenv
from llama_cloud_services import LlamaParse
import openpyxl

load_dotenv()
LLAMA_CLOUD_API_KEY = os.getenv("LLAMA_CLOUD_API_KEY")
if not LLAMA_CLOUD_API_KEY:
    raise RuntimeError("Please set LLAMA_CLOUD_API_KEY in your environment.")

parser = LlamaParse(
    api_key=LLAMA_CLOUD_API_KEY,
    result_type="markdown",
    verbose=True,
    premium_mode=True,
    extract_layout=True,
    adaptive_long_table=True,
)

# -------------------------------------------------------------
# Helper: sanitize sheet name for use as a filename
# -------------------------------------------------------------

def sanitize_filename(name: str) -> str:
    return re.sub(r'[\\/*?:"<>|]', "_", name).strip()

# -------------------------------------------------------------
# Helper: extract markdown from a JobResult object
# -------------------------------------------------------------

def extract_markdown_from_job(job_result) -> str:
    if hasattr(job_result, "text") and job_result.text:
        return job_result.text

    elif hasattr(job_result, "markdown") and job_result.markdown:
        return job_result.markdown

    elif hasattr(job_result, "documents"):
        parts = []
        for i, doc in enumerate(job_result.documents):
            content = doc.text if hasattr(doc, "text") else str(doc)
            parts.append(f"<!-- Page {i + 1} -->\n\n{content}")
        return "\n\n---\n\n".join(parts)

    elif hasattr(job_result, "pages"):
        parts = []
        for i, page in enumerate(job_result.pages):
            content = page.md if hasattr(page, "md") else str(page)
            parts.append(f"<!-- Page {i + 1} -->\n\n{content}")
        return "\n\n---\n\n".join(parts)

    else:
        return str(job_result)

# -------------------------------------------------------------
# Main: parse each sheet separately
# -------------------------------------------------------------

def parse_each_sheet(excel_path: str, out_dir: str):
    os.makedirs(out_dir, exist_ok=True)

    # Load original workbook
    wb_original = openpyxl.load_workbook(excel_path, data_only=True)
    sheet_names = wb_original.sheetnames
    print(f"[info] Found {len(sheet_names)} sheets: {sheet_names}\n")

    results = {}  # sheet_name -> output md path

    for sheet_name in sheet_names:
        print(f"[sheet] Processing: '{sheet_name}'")

        # 1) Create a new workbook with only this sheet
        wb_single = openpyxl.Workbook()
        ws_new = wb_single.active
        ws_new.title = sheet_name

        ws_orig = wb_original[sheet_name]
        for row in ws_orig.iter_rows():
            for cell in row:
                ws_new[cell.coordinate] = cell.value

        # 2) Save temp single-sheet xlsx
        safe_name = sanitize_filename(sheet_name)
        temp_xlsx = os.path.join(out_dir, f"_temp_{safe_name}.xlsx")
        wb_single.save(temp_xlsx)

        # 3) Call LlamaParse on this single-sheet xlsx
        try:
            job_result = parser.parse(temp_xlsx)
            markdown = extract_markdown_from_job(job_result)
        except Exception as e:
            print(f"  [error] LlamaParse failed for sheet '{sheet_name}': {e}")
            markdown = f"<!-- LlamaParse failed for sheet: {sheet_name} -->\n"

        # 4) Save markdown for this sheet
        out_md_path = os.path.join(out_dir, f"sheet_{safe_name}.md")
        with open(out_md_path, "w", encoding="utf-8") as f:
            f.write(f"# Sheet: {sheet_name}\n\n")
            f.write(markdown)

        print(f"  [done] Markdown written to {out_md_path} ({len(markdown)} chars)")
        results[sheet_name] = out_md_path

        # 5) Clean up temp file
        os.remove(temp_xlsx)

    print(f"\n[done] All {len(sheet_names)} sheets processed.")
    print(f"[done] Markdown files saved in: {out_dir}")
    return results

# -------------------------------------------------------------
# Entry point
# -------------------------------------------------------------

if __name__ == "__main__":
    EXCEL_PATH = r"C:\Users\Nemro Neno\Desktop\semesters\Semester 8\LLM\Project\qa_excel.xlsx"
    OUT_DIR    = "sheets_markdown"   # folder where per-sheet .md files will be saved

    parse_each_sheet(EXCEL_PATH, OUT_DIR)
